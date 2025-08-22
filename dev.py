import json
import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

LISTING_URL = "https://www.devnetjobsindia.org/rfp_assignments.aspx"
DETAIL_URL = "https://www.devnetjobsindia.org/JobDescription.aspx?Job_Id={jobid}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

# --------------------------
# Custom "How To Apply" keywords
# --------------------------
HOW_TO_APPLY_KEYWORDS = [
    "Selection Criteria", "Evaluation & Follow-Up", "Application Guidelines", "Eligible Applicants:",
    "Scope of Work:", "Proposal Requirements", "Evaluation Criteria", "Submission Details", "Eligible Entities",
    "How to apply", "Purpose of RFP", "Proposal Guidelines", "Eligibility Criteria", "Application must include:",
    "Eligibility", "Submission of Tender:", "Technical Bid-", "Who Can Apply", "Documents Required", "Expectation:",
    "Eligibility Criterion:", "Submission terms:", "Vendor Qualifications", "To apply",
    "To know about the eligibility criteria:", "The agency's specific responsibilities include –",
    "SELCO Foundation will be responsible for:", "Partner Eligibility Criteria", "Proposal Submission Requirements",
    "Proposal Evaluation Criteria", "Eligibility Criteria for CSOs to be part of the programme:", "Pre-Bid Queries:",
    "Response to Pre-Bid Queries:", "Submission of Bid:", "Applicant Profiles:", "What we like to see in grant applications:",
    "Research that is supported by the SVRI must:", "Successful projects are most often:", "Criteria for funding:",
    "Before you begin to write your proposal, consider that IEF prefers to fund:",
    "As you prepare your budget, these are some items that IEF will not fund:", "Organizational Profile",
    "Selection Process", "Proposal Submission Guidelines", "Terms and Conditions", "Security Deposit:",
    "Facilities and Support Offered under the call for proposal:", "Prospective Consultants should demonstrate:", " Submission of bids","Protocol for Applying",
    "Team Leader background:", "Education:", "Work Experience:","Languages:", "Instructions and Deadlines for Responding",
    "Passing Gifts Private Limited (PGPL)","HIRING OF AN AGENCY FOR DOCUMENTING GOOD PRACTICES IN PROGRAMMING FOR ADOLESCENT GIRLS IN INDIA UNDER THE UNFPA’S UNFPA-SUPPORTED ASTITVA NATIONAL PROJECT.",
    "Application Process",
]

# --------------------------
# Load verticals
# --------------------------
def load_verticals(path="keywords.json"):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("verticals", {})

def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def match_verticals(text: str, verticals: dict) -> list:
    t = text.lower()
    matched = []
    for vertical, kws in verticals.items():
        for kw in kws:
            if kw.lower() in t:
                matched.append(vertical)
                break
    return matched

# --------------------------
# ASP.NET helpers
# --------------------------
def get_hidden_fields(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    fields = {}
    for field in ["__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"]:
        tag = soup.select_one(f"#{field}")
        if tag and tag.has_attr("value"):
            fields[field] = tag["value"]
    return fields

def simulate_postback(session: requests.Session, hidden: dict, event_target: str) -> str:
    payload = {
        "__EVENTTARGET": event_target,
        "__EVENTARGUMENT": "",
    }
    payload.update(hidden)

    resp = session.post(LISTING_URL, data=payload, headers=HEADERS, allow_redirects=True, timeout=30)

    if "JobDescription.aspx?Job_Id=" in resp.url:
        m = re.search(r"JobDescription\.aspx\?Job_Id=(\d+)", resp.url, re.I)
        if m:
            return DETAIL_URL.format(jobid=m.group(1))

    m = re.search(r"JobDescription\.aspx\?Job_Id=(\d+)", resp.text, re.I)
    if m:
        return DETAIL_URL.format(jobid=m.group(1))

    return ""

# --------------------------
# Extractors
# --------------------------
def fetch_detail_page(session: requests.Session, link: str) -> str:
    if not link:
        return ""
    try:
        resp = session.get(link, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        page_text = soup.get_text("\n", strip=True)

        m_start = re.search(r"(?i)job\s*id\s*:\s*\d+", page_text)
        start_idx = m_start.start() if m_start else 0
        m_end = re.search(r"(?i)view\s*similar\s*jobs\s*:?", page_text)
        end_idx = m_end.start() if (m_end and m_end.start() > start_idx) else len(page_text)

        return page_text[start_idx:end_idx].strip()
    except Exception as e:
        print(f"⚠️ Failed to fetch detail page {link}: {e}")
        return ""

def extract_how_to_apply(description: str) -> str:
    if not description:
        return ""
    lines = description.splitlines()
    matched_sections = []
    i = 0
    while i < len(lines):
        line = lines[i]
        for kw in HOW_TO_APPLY_KEYWORDS:
            if kw.lower() in line.lower():
                section = ["• " + line.strip()]
                i += 1
                while i < len(lines) and not any(k.lower() in lines[i].lower() for k in HOW_TO_APPLY_KEYWORDS):
                    section.append(lines[i].strip())
                    i += 1
                matched_sections.append("\n".join(section).strip())
                break
        else:
            i += 1
    return "\n\n".join(matched_sections).strip()

def extract_rows(html: str):
    soup = BeautifulSoup(html, "html.parser")
    return soup.select("tr.gridRow, tr.gridAltRow")

def build_link_from_logo(row) -> str:
    img = row.select_one("img[src*='joblogos/']")
    if not img or not img.has_attr("src"):
        return ""
    m = re.search(r"joblogos/(\d+)", img["src"])
    if not m:
        return ""
    return DETAIL_URL.format(jobid=m.group(1))

def extract_event_target_from_href(href: str) -> str:
    if not href or "javascript:__doPostBack" not in href:
        return ""
    m = re.search(r"__doPostBack\('([^']+)'", href)
    return m.group(1) if m else ""

def extract_assignments(session: requests.Session, html: str, hidden: dict, verticals: dict):
    results = []
    for row in extract_rows(html):
        a_title = row.select_one("a[id*='lnkJobTitle']")
        title = normalize_text(a_title.get_text(strip=True) if a_title else "")

        org = normalize_text((row.select_one("span[id*='lblJobCo']") or {}).get_text(strip=True) if row.select_one("span[id*='lblJobCo']") else "")
        loc_text = normalize_text((row.select_one("span[id*='lblLocation']") or {}).get_text(strip=True) if row.select_one("span[id*='lblLocation']") else "")
        location = normalize_text(re.sub(r"^Location:\s*", "", loc_text, flags=re.I))

        deadline_text = normalize_text((row.select_one("span[id*='lblApplyDate']") or {}).get_text(strip=True) if row.select_one("span[id*='lblApplyDate']") else "")
        deadline = normalize_text(re.sub(r"^Apply by:\s*", "", deadline_text, flags=re.I))

        base_description = " | ".join([p for p in [org, location] if p])
        matched_verticals = match_verticals(f"{title} {base_description}", verticals)
        if not matched_verticals:
            continue

        link = build_link_from_logo(row)
        if not link and a_title:
            event_target = extract_event_target_from_href(a_title.get("href", ""))
            if event_target:
                link = simulate_postback(session, hidden, event_target)
                time.sleep(0.6)

        full_desc = fetch_detail_page(session, link)
        description = f"{base_description}\n\n{full_desc}" if full_desc else base_description
        how_to_apply = extract_how_to_apply(full_desc)

        results.append({
            "Title": title,
            "Description": description,
            "How_To_Apply": how_to_apply,
            "Deadline": deadline,
            "Matched_Verticals": ", ".join(sorted(set(matched_verticals))),
            "Link": link or ""
        })
    return results

# --------------------------
# Excel helpers
# --------------------------
def save_excel_clickable(rows, filename="relevant_rfps.xlsx"):
    df = pd.DataFrame(rows, columns=["Title", "Description", "How_To_Apply", "Deadline", "Matched_Verticals", "Link"])
    df.to_excel(filename, index=False)

    try:
        wb = load_workbook(filename)
        ws = wb.active

        # Clickable links (column F)
        for r in range(2, ws.max_row + 1):
            cell = ws[f"F{r}"]
            url = cell.value
            if isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.style = "Hyperlink"

        # Column widths
        col_widths = {"A": 40, "B": 100, "C": 80, "D": 18, "E": 25, "F": 60}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Row height + alignment
        fixed_height = 40
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = fixed_height
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if c in (2, 3):  # Description + How_To_Apply
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

        wb.save(filename)
    except Exception as e:
        print(f"⚠️ Could not format Excel file: {e}")

# --------------------------
# Deadline Parser
# --------------------------
def parse_deadline(deadline_str: str):
    """Convert deadline string into datetime"""
    if not deadline_str:
        return datetime.max
    for fmt in ("%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(deadline_str.strip(), fmt)
        except ValueError:
            continue
    return datetime.max

def format_deadline(deadline_str: str):
    dt = parse_deadline(deadline_str)
    if dt == datetime.max:
        return deadline_str
    return dt.strftime("%d-%m-%Y")  # numeric month

# --------------------------
# Main
# --------------------------
def main():
    print("Fetching RFP assignments...")
    verticals = load_verticals("keywords.json")

    session = requests.Session()
    resp = session.get(LISTING_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    hidden = get_hidden_fields(resp.text)
    rows = extract_assignments(session, resp.text, hidden, verticals)

    if not rows:
        print("❌ No relevant assignments found with given keywords.")
        return

    today = datetime.today()

    # 🔹 Sort: upcoming deadlines first (earliest first), expired later
    rows.sort(key=lambda r: (
        parse_deadline(r["Deadline"]) < today,  # False (upcoming) first
        parse_deadline(r["Deadline"])           # then by actual date
    ))

    # 🔹 Reformat all deadlines as DD-MM-YYYY
    for r in rows:
        r["Deadline"] = format_deadline(r["Deadline"])

    save_excel_clickable(rows, "devnetjobindiascraper.xlsx")
    print(f"✅ Saved {len(rows)} relevant assignments to devnetjobindiascraper.xlsx (sorted by upcoming deadlines first)")

if __name__ == "__main__":
    main()
